from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Sum, Q
from django.utils import timezone
from .models import Product, StockIn, Sale, Branch, UserProfile
from .forms import StockInForm, SaleForm, ProductForm, BranchForm, UserBranchForm


# ─── Helpers ──────────────────────────────────────────────────────────────────

def get_user_profile(user):
    profile, _ = UserProfile.objects.get_or_create(
        user=user,
        defaults={'role': 'staff', 'branch': Branch.objects.first()}
    )
    return profile


def is_owner(user):
    return get_user_profile(user).is_owner()


def get_branch_scope(request):
    """Returns the branch to scope data to, or None if owner (sees all)."""
    profile = get_user_profile(request.user)
    if profile.is_owner():
        # Owner can switch branch via session or see all
        branch_id = request.session.get('active_branch_id')
        if branch_id:
            try:
                return Branch.objects.get(pk=branch_id)
            except Branch.DoesNotExist:
                request.session.pop('active_branch_id', None)
        return None  # None = all branches
    return profile.branch


def owner_required(view_func):
    """Decorator: only owners can access this view."""
    def wrapper(request, *args, **kwargs):
        if not is_owner(request.user):
            messages.error(request, 'Only the owner can access this page.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


# ─── Auth ─────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        messages.error(request, 'Invalid username or password.')
    return render(request, 'stock/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ─── Branch switcher (owner only) ─────────────────────────────────────────────

@login_required
def switch_branch(request, pk):
    if not is_owner(request.user):
        return redirect('dashboard')
    if pk == 0:
        request.session.pop('active_branch_id', None)
        messages.success(request, 'Now viewing ALL branches.')
    else:
        branch = get_object_or_404(Branch, pk=pk)
        request.session['active_branch_id'] = branch.pk
        messages.success(request, f'Now viewing: {branch.name}')
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


# ─── Dashboard ────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    profile = get_user_profile(request.user)
    branch = get_branch_scope(request)

    sales_qs = Sale.objects.all()
    stockin_qs = StockIn.objects.all()
    if branch:
        sales_qs = sales_qs.filter(branch=branch)
        stockin_qs = stockin_qs.filter(branch=branch)

    total_revenue = sales_qs.aggregate(t=Sum('payment_received'))['t'] or 0
    total_pending = sales_qs.aggregate(t=Sum('payment_pending'))['t'] or 0
    total_purchase_val = sum(s.total_purchase_value for s in stockin_qs)
    products = Product.objects.all()
    out_count = sum(1 for p in products if p.current_stock <= 0)
    low_count = sum(1 for p in products if 0 < p.current_stock <= 5)

    all_branches = Branch.objects.filter(is_active=True) if profile.is_owner() else None
    active_branch_id = request.session.get('active_branch_id')

    context = {
        'profile': profile,
        'branch': branch,
        'products': products,
        'total_revenue': total_revenue,
        'total_pending': total_pending,
        'total_purchase_val': total_purchase_val,
        'out_count': out_count,
        'low_count': low_count,
        'total_sales_count': sales_qs.count(),
        'all_branches': all_branches,
        'active_branch_id': active_branch_id,
    }
    return render(request, 'stock/dashboard.html', context)


# ─── Stock In ─────────────────────────────────────────────────────────────────

@login_required
def stock_in_list(request):
    profile = get_user_profile(request.user)
    branch = get_branch_scope(request)
    q = request.GET.get('q', '')
    entries = StockIn.objects.select_related('product', 'branch').all()
    if branch:
        entries = entries.filter(branch=branch)
    if q:
        entries = entries.filter(Q(product__name__icontains=q) | Q(supplier__icontains=q))
    return render(request, 'stock/stockin_list.html', {
        'entries': entries, 'q': q, 'branch': branch, 'profile': profile
    })


@login_required
def stock_in_add(request):
    profile = get_user_profile(request.user)
    branch = get_branch_scope(request)
    form = StockInForm(request.POST or None, user_branch=branch, is_owner=profile.is_owner())
    if request.method == 'POST' and form.is_valid():
        entry = form.save(commit=False)
        if not profile.is_owner():
            entry.branch = branch
        entry.save()
        product = entry.product
        product.current_stock += entry.quantity_in
        product.save()
        messages.success(request, f'Stock added: {entry.quantity_in} units of {product.name}. New stock: {product.current_stock}')
        return redirect('stockin_list')
    return render(request, 'stock/stockin_form.html', {
        'form': form, 'title': 'Add Stock In', 'profile': profile
    })


@login_required
def stock_in_edit(request, pk):
    profile = get_user_profile(request.user)
    branch = get_branch_scope(request)
    entry = get_object_or_404(StockIn, pk=pk)
    if not profile.is_owner() and entry.branch != branch:
        messages.error(request, 'You can only edit stock for your branch.')
        return redirect('stockin_list')
    old_qty = entry.quantity_in
    form = StockInForm(request.POST or None, instance=entry, user_branch=branch, is_owner=profile.is_owner())
    if request.method == 'POST' and form.is_valid():
        updated = form.save()
        diff = updated.quantity_in - old_qty
        product = updated.product
        product.current_stock += diff
        product.save()
        messages.success(request, f'Stock entry updated. Stock adjusted by {diff:+d}.')
        return redirect('stockin_list')
    return render(request, 'stock/stockin_form.html', {
        'form': form, 'title': 'Edit Stock In', 'profile': profile
    })


@login_required
def stock_in_delete(request, pk):
    profile = get_user_profile(request.user)
    branch = get_branch_scope(request)
    entry = get_object_or_404(StockIn, pk=pk)
    if not profile.is_owner() and entry.branch != branch:
        messages.error(request, 'You can only delete stock for your branch.')
        return redirect('stockin_list')
    if request.method == 'POST':
        product = entry.product
        product.current_stock -= entry.quantity_in
        product.save()
        entry.delete()
        messages.success(request, f'Stock entry deleted. Stock reduced by {entry.quantity_in}.')
        return redirect('stockin_list')
    return render(request, 'stock/confirm_delete.html', {'obj': entry, 'back': 'stockin_list'})


# ─── Sales ────────────────────────────────────────────────────────────────────

@login_required
def sale_list(request):
    profile = get_user_profile(request.user)
    branch = get_branch_scope(request)

    q = request.GET.get('q', '')
    customer_filter = request.GET.get('customer', '')
    product_filter = request.GET.get('product', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    qty_min = request.GET.get('qty_min', '')
    qty_max = request.GET.get('qty_max', '')
    discount_base_filter = request.GET.get('discount_base', '')
    disc_min = request.GET.get('disc_min', '')
    disc_max = request.GET.get('disc_max', '')

    sales = Sale.objects.select_related('product', 'branch').all()
    if branch:
        sales = sales.filter(branch=branch)
    if q:
        sales = sales.filter(Q(customer_name__icontains=q) | Q(product__name__icontains=q))
    if customer_filter:
        sales = sales.filter(customer_name=customer_filter)
    if product_filter:
        sales = sales.filter(product__id=product_filter)
    if date_from:
        sales = sales.filter(date__gte=date_from)
    if date_to:
        sales = sales.filter(date__lte=date_to)
    if qty_min:
        sales = sales.filter(quantity_sold__gte=qty_min)
    if qty_max:
        sales = sales.filter(quantity_sold__lte=qty_max)
    if discount_base_filter:
        sales = sales.filter(discount_base=discount_base_filter)
    if disc_min:
        sales = sales.filter(discount_percent__gte=disc_min)
    if disc_max:
        sales = sales.filter(discount_percent__lte=disc_max)

    all_customers = Sale.objects.values_list('customer_name', flat=True).distinct().order_by('customer_name')
    all_products = Product.objects.all().order_by('name')

    context = {
        'sales': sales, 'profile': profile, 'branch': branch,
        'q': q, 'customer_filter': customer_filter,
        'product_filter': product_filter, 'date_from': date_from, 'date_to': date_to,
        'qty_min': qty_min, 'qty_max': qty_max,
        'discount_base_filter': discount_base_filter,
        'disc_min': disc_min, 'disc_max': disc_max,
        'all_customers': all_customers, 'all_products': all_products,
    }
    return render(request, 'stock/sale_list.html', context)


@login_required
def sale_add(request):
    profile = get_user_profile(request.user)
    branch = get_branch_scope(request)
    form = SaleForm(request.POST or None, user_branch=branch, is_owner=profile.is_owner())
    if request.method == 'POST' and form.is_valid():
        sale = form.save(commit=False)
        if not profile.is_owner():
            sale.branch = branch
        sale.mrp = sale.product.mrp
        sale.selling_price = sale.calculate_selling_price()
        sale.save()
        product = sale.product
        product.current_stock -= sale.quantity_sold
        product.save()
        messages.success(request, f'Sale recorded! {sale.quantity_sold} units of {product.name} sold. Remaining stock: {product.current_stock}')
        return redirect('sale_list')
    return render(request, 'stock/sale_form.html', {
        'form': form, 'title': 'Add Sale',
        'products': Product.objects.all(), 'profile': profile
    })


@login_required
def sale_edit(request, pk):
    profile = get_user_profile(request.user)
    branch = get_branch_scope(request)
    sale = get_object_or_404(Sale, pk=pk)
    if not profile.is_owner() and sale.branch != branch:
        messages.error(request, 'You can only edit sales for your branch.')
        return redirect('sale_list')
    old_qty = sale.quantity_sold
    form = SaleForm(request.POST or None, instance=sale, user_branch=branch, is_owner=profile.is_owner())
    if request.method == 'POST' and form.is_valid():
        updated = form.save(commit=False)
        updated.mrp = updated.product.mrp
        updated.selling_price = updated.calculate_selling_price()
        updated.save()
        diff = updated.quantity_sold - old_qty
        product = updated.product
        product.current_stock -= diff
        product.save()
        messages.success(request, f'Sale updated. Stock adjusted by {-diff:+d}.')
        return redirect('sale_list')
    return render(request, 'stock/sale_form.html', {
        'form': form, 'title': 'Edit Sale',
        'products': Product.objects.all(), 'profile': profile
    })


@login_required
def sale_delete(request, pk):
    profile = get_user_profile(request.user)
    branch = get_branch_scope(request)
    sale = get_object_or_404(Sale, pk=pk)
    if not profile.is_owner() and sale.branch != branch:
        messages.error(request, 'You can only delete sales for your branch.')
        return redirect('sale_list')
    if request.method == 'POST':
        product = sale.product
        product.current_stock += sale.quantity_sold
        product.save()
        sale.delete()
        messages.success(request, f'Sale deleted. {sale.quantity_sold} units returned to stock.')
        return redirect('sale_list')
    return render(request, 'stock/confirm_delete.html', {'obj': sale, 'back': 'sale_list'})


@login_required
def sale_invoice(request, pk):
    sale = get_object_or_404(Sale.objects.select_related('product', 'branch'), pk=pk)
    invoice_number = f"SKN-{sale.date.strftime('%Y%m')}-{sale.pk:04d}"
    return render(request, 'stock/invoice.html', {
        'sale': sale,
        'invoice_number': invoice_number,
        'generated_on': timezone.localdate(),
    })


@login_required
def sale_invoice_bulk(request):
    ids = request.GET.get('ids', '')
    id_list = [int(i) for i in ids.split(',') if i.strip().isdigit()]
    sales = Sale.objects.select_related('product', 'branch').filter(pk__in=id_list)
    grand_total = sum(s.total_sale_value for s in sales)
    grand_received = sum(s.payment_received for s in sales)
    grand_pending = sum(s.payment_pending for s in sales)
    return render(request, 'stock/invoice_bulk.html', {
        'sales': sales,
        'invoice_number': f"SKN-BULK-{timezone.localdate().strftime('%Y%m%d')}",
        'generated_on': timezone.localdate(),
        'grand_total': grand_total,
        'grand_received': grand_received,
        'grand_pending': grand_pending,
    })


# ─── Stock Alerts ─────────────────────────────────────────────────────────────

@login_required
def stock_alert(request):
    products = Product.objects.all()
    out_count = sum(1 for p in products if p.current_stock <= 0)
    low_count = sum(1 for p in products if 0 < p.current_stock <= 5)
    ok_count  = sum(1 for p in products if p.current_stock > 5)
    profile = get_user_profile(request.user)
    return render(request, 'stock/stock_alert.html', {
        'products': products,
        'out_count': out_count, 'low_count': low_count, 'ok_count': ok_count,
        'profile': profile,
    })


# ─── Products ─────────────────────────────────────────────────────────────────

@login_required
def product_list(request):
    profile = get_user_profile(request.user)
    return render(request, 'stock/product_list.html', {
        'products': Product.objects.all(), 'profile': profile
    })


@login_required
def product_add(request):
    profile = get_user_profile(request.user)
    form = ProductForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Product added.')
        return redirect('product_list')
    return render(request, 'stock/product_form.html', {
        'form': form, 'title': 'Add Product', 'profile': profile
    })


@login_required
def product_edit(request, pk):
    profile = get_user_profile(request.user)
    product = get_object_or_404(Product, pk=pk)
    form = ProductForm(request.POST or None, instance=product)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Product updated.')
        return redirect('product_list')
    return render(request, 'stock/product_form.html', {
        'form': form, 'title': 'Edit Product', 'profile': profile
    })


# ─── Branch Management (Owner Only) ───────────────────────────────────────────

@login_required
@owner_required
def branch_list(request):
    branches = Branch.objects.all().prefetch_related('users__user')
    profile = get_user_profile(request.user)
    branch_data = []
    for b in branches:
        sales_count = Sale.objects.filter(branch=b).count()
        stockin_count = StockIn.objects.filter(branch=b).count()
        users = b.users.select_related('user').all()
        branch_data.append({
            'branch': b,
            'sales_count': sales_count,
            'stockin_count': stockin_count,
            'users': users,
        })
    return render(request, 'stock/branch_list.html', {
        'branch_data': branch_data, 'profile': profile
    })


@login_required
@owner_required
def branch_add(request):
    profile = get_user_profile(request.user)
    form = BranchForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f'Branch "{form.cleaned_data["name"]}" created successfully.')
        return redirect('branch_list')
    return render(request, 'stock/branch_form.html', {
        'form': form, 'title': 'Add New Branch', 'profile': profile
    })


@login_required
@owner_required
def branch_edit(request, pk):
    profile = get_user_profile(request.user)
    branch = get_object_or_404(Branch, pk=pk)
    from .forms import BranchForm
    form = BranchForm(request.POST or None, instance=branch)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f'Branch updated.')
        return redirect('branch_list')
    return render(request, 'stock/branch_bank_form.html', {
        'form': form, 'title': f'Edit {branch.name}',
        'branch': branch, 'profile': profile, 'is_owner': True,
    })


@login_required
@owner_required
def branch_delete(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == 'POST':
        branch_name = str(branch)
        branch.delete()
        messages.success(request, f'Branch "{branch_name}" deleted.')
        return redirect('branch_list')
    return render(request, 'stock/confirm_delete.html', {'obj': branch, 'back': 'branch_list'})


@login_required
@owner_required
def branch_assign_user(request, pk):
    profile = get_user_profile(request.user)
    branch = get_object_or_404(Branch, pk=pk)
    from django.contrib.auth.models import User
    form = UserBranchForm(request.POST or None, branch=branch)
    if request.method == 'POST' and form.is_valid():
        user = form.cleaned_data['user']
        role = form.cleaned_data['role']
        up, _ = UserProfile.objects.get_or_create(user=user, defaults={'role': role, 'branch': branch})
        up.branch = branch
        up.role = role
        up.save()
        messages.success(request, f'{user.username} assigned to {branch.name} as {role}.')
        return redirect('branch_list')
    return render(request, 'stock/branch_assign.html', {
        'form': form, 'branch': branch, 'profile': profile
    })


@login_required
@owner_required
def branch_remove_user(request, branch_pk, user_pk):
    from django.contrib.auth.models import User
    branch = get_object_or_404(Branch, pk=branch_pk)
    user = get_object_or_404(User, pk=user_pk)
    if request.method == 'POST':
        UserProfile.objects.filter(user=user, branch=branch).delete()
        messages.success(request, f'{user.username} removed from {branch.name}.')
    return redirect('branch_list')


# ─── Commission Views (Owner Only) ────────────────────────────────────────────

from .models import CommissionSetting, MonthlyCommission
from .forms import CommissionSettingForm, CommissionPaymentForm
from django.db.models import Sum
import calendar
from decimal import Decimal


@login_required
@owner_required
def commission_dashboard(request):
    """Main commission overview — all branches, all months."""
    profile = get_user_profile(request.user)
    branches = Branch.objects.filter(is_active=True).prefetch_related('commission_setting')

    # Totals across all time
    total_commission = MonthlyCommission.objects.aggregate(t=Sum('commission_amount'))['t'] or 0
    total_paid       = MonthlyCommission.objects.aggregate(t=Sum('amount_paid'))['t'] or 0
    total_pending    = sum(
        max(0, m.amount_pending) for m in MonthlyCommission.objects.all()
    )

    # Per-branch summary
    branch_summary = []
    for b in branches:
        setting = CommissionSetting.objects.filter(branch=b).first()
        commissions = MonthlyCommission.objects.filter(branch=b)
        branch_total     = commissions.aggregate(t=Sum('commission_amount'))['t'] or 0
        branch_paid      = commissions.aggregate(t=Sum('amount_paid'))['t'] or 0
        branch_pending   = sum(max(0, m.amount_pending) for m in commissions)
        pending_months   = commissions.filter(status__in=['pending','partial']).count()
        branch_summary.append({
            'branch':         b,
            'setting':        setting,
            'total':          branch_total,
            'paid':           branch_paid,
            'pending':        branch_pending,
            'pending_months': pending_months,
        })

    context = {
        'profile': profile,
        'branch_summary': branch_summary,
        'total_commission': total_commission,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'recent': MonthlyCommission.objects.select_related('branch').all()[:20],
        'today': timezone.localdate(),
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'years': range(timezone.localdate().year - 2, timezone.localdate().year + 1),
    }
    return render(request, 'stock/commission_dashboard.html', context)


@login_required
@owner_required
def commission_setting(request, branch_pk):
    """Owner sets/edits the commission rate for a branch."""
    profile = get_user_profile(request.user)
    branch = get_object_or_404(Branch, pk=branch_pk)
    setting, _ = CommissionSetting.objects.get_or_create(
        branch=branch,
        defaults={'commission_type': 'percentage', 'commission_value': 0}
    )
    form = CommissionSettingForm(request.POST or None, instance=setting)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f'Commission setting saved for {branch.name}.')
        return redirect('commission_dashboard')
    return render(request, 'stock/commission_setting.html', {
        'form': form, 'branch': branch, 'setting': setting, 'profile': profile
    })


@login_required
@owner_required
def commission_generate(request, branch_pk):
    """Generate (or regenerate) commission for a specific month/year."""
    profile = get_user_profile(request.user)
    branch = get_object_or_404(Branch, pk=branch_pk)
    setting = get_object_or_404(CommissionSetting, branch=branch)

    if request.method == 'POST':
        month = int(request.POST.get('month'))
        year  = int(request.POST.get('year'))

        # Calculate total sales for this branch in that month
        sales = Sale.objects.filter(
            branch=branch,
            date__month=month,
            date__year=year
        )
        total_sales = sales.aggregate(t=Sum('selling_price'))['t'] or Decimal('0')
        # Multiply by qty
        total_revenue = Decimal('0')
        for s in sales:
            total_revenue += s.total_sale_value

        commission_breakdown = setting.calculate_commission(total_revenue)

        mc, created = MonthlyCommission.objects.get_or_create(
            branch=branch, month=month, year=year,
            defaults={
                'total_sales':              total_revenue,
                'commission_type':          setting.commission_type,
                'commission_value':         setting.commission_value,
                'base_commission':          commission_breakdown['base_commission'],
                'service_charge_percent':   setting.service_charge_percent,
                'service_charge_amount':    commission_breakdown['service_charge'],
                'gross_commission':         commission_breakdown['gross_commission'],
                'tds_percent':              setting.tds_percent,
                'tds_amount':               commission_breakdown['tds_amount'],
                'commission_amount':        commission_breakdown['net_commission'],
                'status':                   'pending',
            }
        )
        if not created:
            mc.total_sales             = total_revenue
            mc.commission_type         = setting.commission_type
            mc.commission_value        = setting.commission_value
            mc.base_commission         = commission_breakdown['base_commission']
            mc.service_charge_percent  = setting.service_charge_percent
            mc.service_charge_amount   = commission_breakdown['service_charge']
            mc.gross_commission        = commission_breakdown['gross_commission']
            mc.tds_percent             = setting.tds_percent
            mc.tds_amount              = commission_breakdown['tds_amount']
            mc.commission_amount       = commission_breakdown['net_commission']
            mc.save()
            messages.success(request, f'Commission regenerated for {branch.name} — {calendar.month_name[month]} {year}.')
        else:
            messages.success(request, f'Commission generated for {branch.name} — {calendar.month_name[month]} {year}: ₹{commission_breakdown["net_commission"]:.2f}')
        return redirect('commission_branch_detail', branch_pk=branch_pk)

    # GET — show form to pick month/year
    today = timezone.localdate()
    return render(request, 'stock/commission_generate.html', {
        'branch': branch, 'setting': setting, 'profile': profile,
        'today': today,
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'years': range(today.year - 2, today.year + 1),
    })


@login_required
@owner_required
def commission_generate_all(request):
    """Generate commission for ALL branches for a given month/year."""
    if request.method == 'POST':
        month = int(request.POST.get('month'))
        year  = int(request.POST.get('year'))
        count = 0
        for branch in Branch.objects.filter(is_active=True):
            setting = CommissionSetting.objects.filter(branch=branch).first()
            if not setting or setting.commission_value == 0:
                continue
            sales = Sale.objects.filter(branch=branch, date__month=month, date__year=year)
            total_revenue = sum(s.total_sale_value for s in sales)
            breakdown = setting.calculate_commission(total_revenue)

            mc, created = MonthlyCommission.objects.get_or_create(
                branch=branch, month=month, year=year,
                defaults={
                    'total_sales':             total_revenue,
                    'commission_type':         setting.commission_type,
                    'commission_value':        setting.commission_value,
                    'base_commission':         breakdown['base_commission'],
                    'service_charge_percent':  setting.service_charge_percent,
                    'service_charge_amount':   breakdown['service_charge'],
                    'gross_commission':        breakdown['gross_commission'],
                    'tds_percent':             setting.tds_percent,
                    'tds_amount':              breakdown['tds_amount'],
                    'commission_amount':       breakdown['net_commission'],
                    'status':                  'pending',
                }
            )
            if not created:
                mc.total_sales             = total_revenue
                mc.commission_type         = setting.commission_type
                mc.commission_value        = setting.commission_value
                mc.base_commission         = breakdown['base_commission']
                mc.service_charge_percent  = setting.service_charge_percent
                mc.service_charge_amount   = breakdown['service_charge']
                mc.gross_commission        = breakdown['gross_commission']
                mc.tds_percent             = setting.tds_percent
                mc.tds_amount              = breakdown['tds_amount']
                mc.commission_amount       = breakdown['net_commission']
                mc.save()
            count += 1
        messages.success(request, f'Commission generated for {count} branches — {calendar.month_name[month]} {year}.')
    return redirect('commission_dashboard')


@login_required
@owner_required
def commission_branch_detail(request, branch_pk):
    """All monthly commissions for a specific branch — owner view."""
    profile = get_user_profile(request.user)
    branch = get_object_or_404(Branch, pk=branch_pk)
    setting = CommissionSetting.objects.filter(branch=branch).first()
    commissions = MonthlyCommission.objects.filter(branch=branch).order_by('-year', '-month')
    total      = commissions.aggregate(t=Sum('commission_amount'))['t'] or 0
    total_paid = commissions.aggregate(t=Sum('amount_paid'))['t'] or 0
    total_pend = sum(max(0, m.amount_pending) for m in commissions)
    today      = timezone.localdate()
    return render(request, 'stock/commission_branch_detail.html', {
        'branch': branch, 'setting': setting,
        'commissions': commissions,
        'total': total, 'total_paid': total_paid, 'total_pending': total_pend,
        'profile': profile, 'today': today,
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'years': range(today.year - 2, today.year + 1),
    })


@login_required
@owner_required
def commission_pay(request, pk):
    """Mark a monthly commission as paid / partial."""
    profile = get_user_profile(request.user)
    mc = get_object_or_404(MonthlyCommission, pk=pk)
    form = CommissionPaymentForm(request.POST or None, instance=mc)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, f'Payment updated for {mc.branch.name} — {mc.month_name} {mc.year}.')
        return redirect('commission_branch_detail', branch_pk=mc.branch.pk)
    return render(request, 'stock/commission_pay.html', {
        'form': form, 'mc': mc, 'profile': profile
    })


@login_required
@owner_required
def commission_export_excel(request):
    """Export monthly commission for all branches to Excel."""
    month = int(request.GET.get('month', timezone.localdate().month))
    year  = int(request.GET.get('year',  timezone.localdate().year))

    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Commission {calendar.month_name[month]} {year}"

    # ── Styles ──────────────────────────────────────────────
    green_fill  = PatternFill("solid", fgColor="2D6A4F")
    head_fill   = PatternFill("solid", fgColor="E8F4EF")
    alt_fill    = PatternFill("solid", fgColor="F7FAF8")
    warn_fill   = PatternFill("solid", fgColor="FEF3C7")
    paid_fill   = PatternFill("solid", fgColor="F0FDF4")
    pend_fill   = PatternFill("solid", fgColor="FDECEA")

    white_bold  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    dark_bold   = Font(name='Calibri', bold=True, color='1A1A1A', size=10)
    dark_normal = Font(name='Calibri', color='1A1A1A', size=10)
    red_bold    = Font(name='Calibri', bold=True, color='C0392B', size=10)
    green_bold  = Font(name='Calibri', bold=True, color='16A34A', size=10)

    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    rupee_fmt = '₹#,##0.00'
    pct_fmt   = '0.00"%"'

    # ── Title Block ─────────────────────────────────────────
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = f"SKINOVATE — Commission Report  |  {calendar.month_name[month]} {year}"
    title_cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
    title_cell.fill = green_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:N2')
    sub_cell = ws['A2']
    sub_cell.value = f"Generated on: {timezone.localdate()}   |   All amounts in INR (₹)"
    sub_cell.font = Font(name='Calibri', color='6B7280', size=9)
    sub_cell.alignment = center
    ws.row_dimensions[2].height = 16

    ws.append([])  # blank row 3

    # ── Header Row ──────────────────────────────────────────
    headers = [
        'Sr.', 'Branch Name', 'Location', 'Type',
        'Rate Type', 'Rate Value',
        'Total Sales (₹)',
        'Base Commission (₹)',
        'Service Charge %', 'Service Charge (₹)',
        'Gross Commission (₹)',
        'TDS %', 'TDS Amount (₹)',
        'Net Commission (₹)',
        'Amount Paid (₹)', 'Amount Pending (₹)',
        'Payment Date', 'Status', 'Notes',
        # Bank detail columns
        'Email', 'Mobile',
        'Bank Name', 'Account Holder',
        'Account Number', 'IFSC Code',
        'UPI ID', 'GPay Mobile', 'PayTM Mobile', 'PhonePe Mobile',
        'Payment Link',
    ]
    ws.append(headers)
    header_row = ws.max_row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        # Bank columns (20 onwards) get a different header colour
        cell.fill  = PatternFill("solid", fgColor="D1FAE5") if col >= 20 else head_fill
        cell.font  = Font(name='Calibri', bold=True,
                          color='065F46' if col >= 20 else '1A1A1A', size=10)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    # ── Data Rows ────────────────────────────────────────────
    commissions = MonthlyCommission.objects.filter(
        month=month, year=year
    ).select_related('branch').order_by('branch__location', 'branch__name')

    grand = {
        'sales': 0, 'base': 0, 'sc': 0, 'gross': 0,
        'tds': 0, 'net': 0, 'paid': 0, 'pending': 0
    }

    for sr, mc in enumerate(commissions, 1):
        is_alt    = (sr % 2 == 0)
        row_fill  = alt_fill if is_alt else PatternFill()
        bank_row_fill = PatternFill("solid", fgColor="F0FDF4") if not is_alt else PatternFill("solid", fgColor="DCFCE7")

        pending = float(mc.amount_pending)
        status_fill = paid_fill if mc.status == 'paid' else (
            warn_fill if mc.status == 'partial' else pend_fill)

        b = mc.branch
        # Mask account number — show last 4 only
        masked_acc = ''
        if b.account_number:
            masked_acc = ('X' * (len(b.account_number) - 4) + b.account_number[-4:]) if len(b.account_number) > 4 else b.account_number

        row = [
            sr,
            b.name,
            b.location,
            b.commission_setting.get_branch_type_display() if hasattr(b, 'commission_setting') and CommissionSetting.objects.filter(branch=b).exists() else '—',
            'Percentage' if mc.commission_type == 'percentage' else 'Fixed',
            float(mc.commission_value),
            float(mc.total_sales),
            float(mc.base_commission),
            float(mc.service_charge_percent),
            float(mc.service_charge_amount),
            float(mc.gross_commission),
            float(mc.tds_percent),
            float(mc.tds_amount),
            float(mc.commission_amount),
            float(mc.amount_paid),
            pending,
            mc.payment_date.strftime('%d-%b-%Y') if mc.payment_date else '—',
            mc.get_status_display(),
            mc.notes or '',
            # Bank details
            b.email or '—',
            b.mobile or '—',
            b.bank_name or '—',
            b.account_holder or '—',
            masked_acc or '—',
            b.ifsc_code or '—',
            b.upi_id or '—',
            b.gpay_mobile or '—',
            b.paytm_mobile or '—',
            b.phonepe_mobile or '—',
            b.payment_link or '—',
        ]
        ws.append(row)
        r = ws.max_row
        ws.row_dimensions[r].height = 18

        for col in range(1, len(row) + 1):
            c = ws.cell(row=r, column=col)
            c.border = border
            if col >= 20:
                c.fill = bank_row_fill
                c.font = Font(name='Calibri', color='065F46', size=10)
                c.alignment = left
            else:
                c.alignment = center if col in (1, 4, 5, 17, 18) else (right if col >= 6 else left)
                if is_alt:
                    c.fill = row_fill

        # Money format (cols 7–16)
        for col in range(7, 17):
            ws.cell(row=r, column=col).number_format = rupee_fmt
        # Percent cols
        ws.cell(row=r, column=9).number_format  = pct_fmt
        ws.cell(row=r, column=12).number_format = pct_fmt

        # Colour net, paid, pending
        ws.cell(row=r, column=14).font = dark_bold
        ws.cell(row=r, column=15).font = green_bold
        ws.cell(row=r, column=16).font = red_bold if pending > 0 else dark_normal

        # Status colour
        ws.cell(row=r, column=18).fill = status_fill
        ws.cell(row=r, column=18).font = dark_bold

        grand['sales']   += float(mc.total_sales)
        grand['base']    += float(mc.base_commission)
        grand['sc']      += float(mc.service_charge_amount)
        grand['gross']   += float(mc.gross_commission)
        grand['tds']     += float(mc.tds_amount)
        grand['net']     += float(mc.commission_amount)
        grand['paid']    += float(mc.amount_paid)
        grand['pending'] += max(0, pending)

    # ── Grand Total Row ──────────────────────────────────────
    ws.append([])
    ws.append([
        '', 'GRAND TOTAL', '', '', '', '',
        grand['sales'], grand['base'], '', grand['sc'],
        grand['gross'], '', grand['tds'],
        grand['net'], grand['paid'], grand['pending'],
        '', '', '', '', '', '', '', '', '', '', '', '', '', '',
    ])
    r = ws.max_row
    ws.row_dimensions[r].height = 22
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=r, column=col)
        c.fill   = green_fill
        c.font   = white_bold
        c.border = border
        c.alignment = right if col >= 7 else center
    ws.cell(row=r, column=2).alignment = left
    for col in [7, 8, 10, 11, 13, 14, 15, 16]:
        ws.cell(row=r, column=col).number_format = rupee_fmt

    # ── Summary Sheet ────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 22

    def s2(label, value, fmt=None, font=None):
        ws2.append([label, value])
        r2 = ws2.max_row
        la = ws2.cell(row=r2, column=1)
        va = ws2.cell(row=r2, column=2)
        la.font = dark_bold
        va.font = font or dark_normal
        va.alignment = right
        if fmt:
            va.number_format = fmt
        la.border = border
        va.border = border

    ws2.append(['SKINOVATE — Commission Summary'])
    ws2.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=13, color='2D6A4F')
    ws2.append([f'{calendar.month_name[month]} {year}'])
    ws2.cell(row=2, column=1).font = Font(name='Calibri', size=11, color='6B7280')
    ws2.append([])

    s2('Total Branches',          commissions.count())
    s2('Total Sales',             grand['sales'],   rupee_fmt, dark_bold)
    s2('Base Commission',         grand['base'],    rupee_fmt)
    s2('Total Service Charge',    grand['sc'],      rupee_fmt)
    s2('Gross Commission',        grand['gross'],   rupee_fmt, dark_bold)
    s2('Total TDS Deducted',      grand['tds'],     rupee_fmt, Font(name='Calibri', bold=True, color='C0392B', size=10))
    s2('Net Commission Payable',  grand['net'],     rupee_fmt, Font(name='Calibri', bold=True, color='2D6A4F', size=11))
    ws2.append([])
    s2('Amount Paid',             grand['paid'],    rupee_fmt, Font(name='Calibri', bold=True, color='16A34A', size=10))
    s2('Amount Pending',          grand['pending'], rupee_fmt, Font(name='Calibri', bold=True, color='C0392B', size=10))

    # ── Bank Details Sheet ────────────────────────────────────
    ws3 = wb.create_sheet("Bank Details")
    bank_headers = [
        'Sr.', 'Branch Name', 'Location', 'Type',
        'Email', 'Mobile',
        'Bank Name', 'Account Holder', 'Account Number (Masked)', 'IFSC Code',
        'UPI ID', 'GPay Mobile', 'PayTM Mobile', 'PhonePe Mobile',
        'Payment Link', 'Last Updated', 'Updated By',
    ]
    ws3.append(['SKINOVATE — Branch Bank Details'])
    ws3.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=13, color='2D6A4F')
    ws3.cell(row=1, column=1).fill = green_fill
    ws3.merge_cells(f'A1:{get_column_letter(len(bank_headers))}1')
    ws3.row_dimensions[1].height = 26

    ws3.append([f'{calendar.month_name[month]} {year} — Exported: {timezone.localdate()}'])
    ws3.cell(row=2, column=1).font = Font(name='Calibri', size=10, color='6B7280')
    ws3.merge_cells(f'A2:{get_column_letter(len(bank_headers))}2')
    ws3.row_dimensions[2].height = 16

    ws3.append([])

    ws3.append(bank_headers)
    bh_row = ws3.max_row
    for col, h in enumerate(bank_headers, 1):
        c = ws3.cell(row=bh_row, column=col, value=h)
        c.font   = Font(name='Calibri', bold=True, color='065F46', size=10)
        c.fill   = PatternFill("solid", fgColor="D1FAE5")
        c.alignment = center
        c.border = border
    ws3.row_dimensions[bh_row].height = 20

    all_branches = Branch.objects.filter(is_active=True).select_related('bank_details_updated_by').order_by('location','name')
    for sr, b in enumerate(all_branches, 1):
        masked_acc = ''
        if b.account_number:
            masked_acc = ('X' * (len(b.account_number) - 4) + b.account_number[-4:]) if len(b.account_number) > 4 else b.account_number
        btype = '—'
        try:
            cs_obj = CommissionSetting.objects.get(branch=b)
            btype = cs_obj.get_branch_type_display()
        except CommissionSetting.DoesNotExist:
            pass

        bank_row = [
            sr,
            b.name,
            b.location,
            btype,
            b.email or '—',
            b.mobile or '—',
            b.bank_name or '—',
            b.account_holder or '—',
            masked_acc or '—',
            b.ifsc_code or '—',
            b.upi_id or '—',
            b.gpay_mobile or '—',
            b.paytm_mobile or '—',
            b.phonepe_mobile or '—',
            b.payment_link or '—',
            b.bank_details_updated_at.strftime('%d-%b-%Y %H:%M') if b.bank_details_updated_at else 'Never',
            b.bank_details_updated_by.username if b.bank_details_updated_by else '—',
        ]
        ws3.append(bank_row)
        br = ws3.max_row
        ws3.row_dimensions[br].height = 17
        is_alt3 = (sr % 2 == 0)
        for col in range(1, len(bank_row) + 1):
            c = ws3.cell(row=br, column=col)
            c.border = border
            c.alignment = center if col in (1,) else left
            c.font = dark_normal
            if is_alt3:
                c.fill = PatternFill("solid", fgColor="F0FDF4")
        # Highlight missing bank details
        if not b.bank_name:
            for col in range(7, 15):
                ws3.cell(row=br, column=col).fill = PatternFill("solid", fgColor="FEF3C7")
                ws3.cell(row=br, column=col).font = Font(name='Calibri', color='92400E', size=10)

    # Bank sheet column widths
    bank_widths = [5, 24, 14, 14, 24, 16, 20, 24, 24, 14, 22, 16, 16, 16, 30, 18, 14]
    for i, w in enumerate(bank_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = 'B5'

    # ── Column widths on main sheet ──────────────────────────
    col_widths = [5, 24, 14, 14, 13, 12, 18, 20, 16, 18, 20, 10, 16, 20, 18, 18, 15, 12, 24,
                  24, 16, 20, 24, 20, 14, 22, 16, 16, 16, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes ─────────────────────────────────────────
    ws.freeze_panes = 'B5'

    # ── HTTP response ─────────────────────────────────────────
    filename = f"Skinovate_Commission_{calendar.month_name[month]}_{year}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@owner_required
def commission_payout_single(request, pk):
    """Send payout notification for one MonthlyCommission."""
    from .payout_service import send_payout_notification
    from .models import PayoutLog
    mc = get_object_or_404(MonthlyCommission.objects.select_related('branch'), pk=pk)

    if request.method == 'POST':
        log = send_payout_notification(mc, request.user)
        if 'failed' not in log.email_status and 'failed' not in log.sms_status:
            messages.success(request, f'✅ Payout notification sent to {mc.branch.name} — Email: {log.email_status} | SMS: {log.sms_status}')
        else:
            messages.error(request, f'⚠️ Partial send — Email: {log.email_status} | SMS: {log.sms_status}. Check branch contact details.')
        return redirect(request.POST.get('next', 'commission_dashboard'))

    # GET — show confirmation page
    profile = get_user_profile(request.user)
    payout_logs = PayoutLog.objects.filter(monthly_commission=mc).order_by('-sent_at')[:5]
    return render(request, 'stock/payout_confirm.html', {
        'mc': mc, 'profile': profile, 'payout_logs': payout_logs,
    })


@login_required
@owner_required
def commission_payout_all(request):
    """Send payout notifications for ALL pending/partial commissions in a month."""
    from .payout_service import send_payout_notification
    if request.method == 'POST':
        month = int(request.POST.get('month', timezone.localdate().month))
        year  = int(request.POST.get('year',  timezone.localdate().year))
        mcs   = MonthlyCommission.objects.filter(
            month=month, year=year, status__in=['pending', 'partial']
        ).select_related('branch')

        sent_ok, sent_fail = 0, 0
        for mc in mcs:
            log = send_payout_notification(mc, request.user)
            if 'failed' in log.email_status or 'failed' in log.sms_status:
                sent_fail += 1
            else:
                sent_ok += 1

        msg = f'Payout notifications sent for {calendar.month_name[month]} {year}: ✅ {sent_ok} success'
        if sent_fail:
            msg += f', ❌ {sent_fail} failed (check branch contact details)'
        messages.success(request, msg)
    return redirect('commission_dashboard')


@login_required
@owner_required
def payout_log_list(request):
    """View all payout logs."""
    from .models import PayoutLog
    profile = get_user_profile(request.user)
    logs = PayoutLog.objects.select_related(
        'monthly_commission__branch', 'sent_by'
    ).all()[:100]
    return render(request, 'stock/payout_log_list.html', {
        'logs': logs, 'profile': profile,
    })


@login_required
@owner_required
def branch_bank_details(request, pk):
    """Owner views/edits full bank + payment details for any branch."""
    from .forms import BranchForm
    profile = get_user_profile(request.user)
    branch  = get_object_or_404(Branch, pk=pk)
    form    = BranchForm(request.POST or None, instance=branch)
    if request.method == 'POST' and form.is_valid():
        b = form.save(commit=False)
        b.bank_details_updated_at = timezone.now()
        b.bank_details_updated_by = request.user
        b.save()
        messages.success(request, f'Bank & payment details updated for {branch.name}.')
        return redirect('branch_list')
    return render(request, 'stock/branch_bank_form.html', {
        'form': form, 'branch': branch, 'profile': profile, 'is_owner': True,
    })


@login_required
def branch_my_details(request):
    """Branch staff edit their own branch's bank & payment details."""
    profile = get_user_profile(request.user)
    if profile.is_owner():
        return redirect('branch_list')
    branch = profile.branch
    if not branch:
        messages.error(request, 'No branch assigned to your account.')
        return redirect('dashboard')
    from .forms import BranchStaffBankForm
    form = BranchStaffBankForm(request.POST or None, instance=branch)
    if request.method == 'POST' and form.is_valid():
        b = form.save(commit=False)
        b.bank_details_updated_at = timezone.now()
        b.bank_details_updated_by = request.user
        b.save()
        messages.success(request, 'Your branch bank & payment details have been updated. Owner can now see these changes.')
        return redirect('commission_my_view')
    return render(request, 'stock/branch_my_bank_form.html', {
        'form': form, 'branch': branch, 'profile': profile,
    })


@login_required
def commission_my_view(request):
    """Branch/franchise staff see their own commission history."""
    profile = get_user_profile(request.user)
    if profile.is_owner():
        return redirect('commission_dashboard')
    branch = profile.branch
    if not branch:
        messages.error(request, 'No branch assigned to your account.')
        return redirect('dashboard')
    commissions = MonthlyCommission.objects.filter(branch=branch).order_by('-year', '-month')
    setting = CommissionSetting.objects.filter(branch=branch).first()
    total      = commissions.aggregate(t=Sum('commission_amount'))['t'] or 0
    total_paid = commissions.aggregate(t=Sum('amount_paid'))['t'] or 0
    total_pend = sum(max(0, m.amount_pending) for m in commissions)
    return render(request, 'stock/commission_my_view.html', {
        'branch': branch, 'setting': setting,
        'commissions': commissions,
        'total': total, 'total_paid': total_paid, 'total_pending': total_pend,
        'profile': profile,
    })

# ── One-time Setup View ───────────────────────────────────────────────────────
from django.http import HttpResponse

def run_setup(request):
    secret = request.GET.get('key', '')
    if secret != 'skinovate-setup-2026':
        return HttpResponse('Access denied.', status=403)

    from django.contrib.auth.models import User
    from stock.models import Branch, UserProfile, Product
    from decimal import Decimal
    import io
    from django.core.management import call_command
    log = []

    out = io.StringIO()
    call_command('migrate', '--run-syncdb', stdout=out, verbosity=0)
    log.append('Migrations applied')

    for uname, email, pwd, is_super in [
        ('owner', 'owner@skinovate.com', 'owner@123', True),
        ('admin', 'admin@skinovate.com', 'admin123',  True),
    ]:
        if not User.objects.filter(username=uname).exists():
            u = User.objects.create_superuser(uname, email, pwd)
        else:
            u = User.objects.get(username=uname)
            u.set_password(pwd); u.save()
        UserProfile.objects.update_or_create(user=u, defaults={'role': 'owner', 'branch': None})
    log.append('Owner users created: owner/owner@123 and admin/admin123')

    branches_data = [
        ("Nerul Branch 1",  "Nerul",  "Shop 1, Nerul Plaza"),
        ("Nerul Branch 2",  "Nerul",  "Shop 7, Sector 19"),
        ("Nerul Branch 3",  "Nerul",  "Shop 3, Palm Beach Road"),
        ("Thane Branch 1",  "Thane",  "Shop 12, Viviana Mall"),
        ("Thane Branch 2",  "Thane",  "Shop 5, Kapurbawdi"),
        ("Panvel Branch 1", "Panvel", "Shop 2, New Panvel East"),
        ("Panvel Branch 2", "Panvel", "Shop 8, Kamothe"),
        ("Panvel Branch 3", "Panvel", "Shop 4, Kharghar"),
        ("Panvel Branch 4", "Panvel", "Shop 11, Ulwe"),
        ("Panvel Branch 5", "Panvel", "Shop 6, Kalamboli"),
    ]
    for name, loc, addr in branches_data:
        Branch.objects.get_or_create(name=name, defaults={'location': loc, 'address': addr})
    log.append(f'{Branch.objects.count()} branches created')

    branch_users = [
        ('nerul1','nerul1@123'), ('nerul2','nerul2@123'), ('nerul3','nerul3@123'),
        ('thane1','thane1@123'), ('thane2','thane2@123'),
        ('panvel1','panvel1@123'), ('panvel2','panvel2@123'), ('panvel3','panvel3@123'),
        ('panvel4','panvel4@123'), ('panvel5','panvel5@123'),
    ]
    branches = list(Branch.objects.order_by('pk'))
    for i, (uname, pwd) in enumerate(branch_users):
        branch = branches[i] if i < len(branches) else branches[0]
        if not User.objects.filter(username=uname).exists():
            u = User.objects.create_user(uname, f'{uname}@skinovate.com', pwd)
        else:
            u = User.objects.get(username=uname)
            u.set_password(pwd); u.save()
        UserProfile.objects.update_or_create(user=u, defaults={'role': 'staff', 'branch': branch})
        log.append(f'  {uname}/{pwd} -> {branch.name}')

    products_data = [
        ("Skin Brightening Face Wash", 555, 389),
        ("Sunshield Sunscreen", 779, 545),
        ("Radiance Night Serum", 1499, 1049),
        ("Tab Glutathione Glow Shots", 1499, 1049),
        ("Rice Water Serum", 897, 628),
        ("Gentle Cleanser", 649, 454),
        ("Anti-acne Face Wash", 549, 384),
        ("Skinovate Mosturizer", 751, 526),
        ("Skin Brightening Night Cream", 595, 417),
        ("Hair Cocktail Tablets", 890, 623),
        ("Tab Immunity booster shots", 749, 524),
        ("Hair strengthening Shampoo", 888, 622),
        ("Hair Conditioner", 892, 624),
        ("Hair Oil", 555, 389),
    ]
    for name, mrp, dp in products_data:
        Product.objects.get_or_create(name=name, defaults={
            'mrp': Decimal(str(mrp)), 'dp': Decimal(str(dp)), 'current_stock': 0,
        })
    log.append(f'{Product.objects.count()} products created')
    log.append('SETUP COMPLETE! Login: owner / owner@123')

    return HttpResponse('\n'.join(log), content_type='text/plain')